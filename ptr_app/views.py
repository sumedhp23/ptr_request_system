from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.cache import never_cache
from django.http import JsonResponse
from .forms import PTRRequestForm
from .models import (
    PTRRequest, Department, Customer, Project, PartMaster,
    MachineModel, TypeOfMachine,
)


def is_manager(user):
    return user.is_superuser or user.groups.filter(name='tooling_manager').exists()

def is_designer_or_admin(user):
    return user.is_superuser or user.groups.filter(name='designer').exists()


# ═══════════════════════════ PTR Request ═══════════════════════════


def _generate_ptr_number(unit):
    """Generate PTR number in format: UNIT/FY/serial e.g. ASMIPL/26-27/01
    Serial increments by 1 for every PTR submitted from 1st April of the FY."""
    from datetime import datetime
    now = datetime.now()
    # Financial year: April-March
    if now.month >= 4:
        fy_start = now.year % 100
        fy_end = (now.year + 1) % 100
        fy_start_date = datetime(now.year, 4, 1)
    else:
        fy_start = (now.year - 1) % 100
        fy_end = now.year % 100
        fy_start_date = datetime(now.year - 1, 4, 1)
    
    fy = f"{fy_start:02d}-{fy_end:02d}"
    prefix = f"{unit}/{fy}/"

    # Find the highest serial already used in PTRRequest for this prefix
    max_serial = 0
    existing_entries = PTRRequest.objects.filter(ptrno__startswith=prefix)
    import re
    for entry in existing_entries:
        # Extract the serial number from "UNIT/FY/NN"
        match = re.search(r'/(\d+)$', entry.ptrno)
        if match:
            serial = int(match.group(1))
            if serial > max_serial:
                max_serial = serial

    next_serial = max_serial + 1
    return f"{prefix}{next_serial:02d}"


@login_required
@never_cache
def ptr_request_view(request):
    """Render and process the PTR / TCMR request form."""
    if request.method == "POST":
        form = PTRRequestForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            ptr = PTRRequest(
                unit=cd["unit"],
                deptid=cd["deptid"],
                partid=str(cd["partid"]),
                partcode=cd.get("partcode", ""),
                partname=cd.get("partname", ""),
                customerid=cd.get("customerid", 0),
                projectid=cd.get("projectid", 0),
                process=cd["process"],
                drawing=cd["drawing"],
                dotr=cd["dotr"],
                ncsno=cd.get("ncsno", ""),
                bqty=cd.get("bqty", ""),
                mfg=cd.get("mfg", ""),
                requestremarks=cd.get("requestremarks", ""),
                mmodel=cd.get("mmodel", ""),
                typemachine=cd.get("typemachine", ""),
                status=1,  # PTR filled = status 1
                createdby=request.user.username,
                istcmr=cd.get("istcmr", False),
                req_path=cd.get("req_path", ""),
                earlydatereason=cd.get("earlydatereason", ""),
            )
            ptr.save()
            messages.success(
                request,
                f"PTR Request PTR-{ptr.pk:04d} submitted successfully!",
            )
            return redirect("ptr_success")
    else:
        form = PTRRequestForm()

    from .models import TblRevision, DesignerEntry
    rev_ptrs = TblRevision.objects.filter(isapproved=True).values_list('ptr', flat=True)
    tinos = DesignerEntry.objects.filter(ptr__in=rev_ptrs, TINo__isnull=False).exclude(TINo="").values_list('TINo', flat=True).distinct()
    tinos = sorted(list(tinos))

    return render(request, "ptr_app/ptr_request.html", {
        "form": form,
        "tcmr_tinos": tinos
    })


@login_required
@never_cache
def ptr_success_view(request):
    """Show submission confirmation."""
    return render(request, "ptr_app/ptr_success.html")


# ═══════════════════════════ AJAX API Endpoints ═══════════════════════════


@login_required
def api_departments(request):
    """GET /api/departments/?unit=ASMIPL → JSON list of departments."""
    unit = request.GET.get("unit", "")
    departments = Department.objects.filter(unit=unit).values("dID", "DeptName")
    return JsonResponse(list(departments), safe=False)


@login_required
def api_parts(request):
    """GET /api/parts/?unit=ASMIPL → JSON list of parts filtered by unit."""
    unit = request.GET.get("unit", "")
    parts = PartMaster.objects.filter(unit=unit).values("PartID", "PartName", "PartDesc")
    return JsonResponse(list(parts), safe=False)


@login_required
def api_part_detail(request):
    """GET /api/part-detail/?part_id=5 → JSON with part + project + customer info."""
    part_id = request.GET.get("part_id", "")
    try:
        part = PartMaster.objects.select_related("PrjID", "PrjID__custid").get(PartID=part_id)
        project = part.PrjID
        customer = project.custid
        data = {
            "partid": part.PartID,
            "partcode": part.PartName,
            "partname": part.PartDesc or "",
            "project_name": project.prjname or "",
            "project_id": project.PID,
            "customer_name": customer.CustomerName,
            "customer_id": customer.CustID,
            "fg_part_name": part.PartDesc or "",  # placeholder until real mapping known
        }
        return JsonResponse(data)
    except PartMaster.DoesNotExist:
        return JsonResponse({"error": "Part not found"}, status=404)


@login_required
def api_machines(request):
    """GET /api/machines/?unit=ASMIPL → JSON with machine models + types for that unit."""
    unit = request.GET.get("unit", "")
    models_qs = MachineModel.objects.filter(unit=unit).values("id", "model")
    types_qs = TypeOfMachine.objects.filter(unit=unit).values("id", "type_of_machine")
    return JsonResponse({
        "machine_models": list(models_qs),
        "machine_types": list(types_qs),
    })


@login_required
def api_tcmr_autofill_data(request):
    """GET /api/tcmr-autofill/?tino=X → JSON with PTR old values + revision details."""
    from .models import DesignerEntry, TblRevision
    tino = request.GET.get("tino", "")
    if not tino:
        return JsonResponse({"error": "No TINO provided"}, status=400)
        
    # Get the Designer Entry which has the TINO. We only care about fully verified + approved revs
    d_entry = DesignerEntry.objects.filter(TINo=tino, ptr__status=7).first()
    if not d_entry:
        return JsonResponse({"error": "No verified PTR found for this TINO"}, status=404)
        
    original_ptr = d_entry.ptr
    # Find the approved obsolete request (TblRevision) for this PTR
    rev = TblRevision.objects.filter(ptr=original_ptr, isapproved=True).order_by('-Vdate').first()
    
    if not rev:
        return JsonResponse({"error": "No approved obsolete request found for this TINO"}, status=404)
        
    # We found the original PTR and the approved revision. Build autofill data!
    data = {
        "unit": original_ptr.unit,
        "deptid": original_ptr.deptid,
        "partid": original_ptr.partid,
        "partcode": original_ptr.partcode,
        "partname": original_ptr.partname,
        # We don't have customer_display directly on PTRRequest, but the frontend api-part fills it. We just need to give `partid`
        # and the frontend can re-trigger part lookup, OR we can pass it here. Let's pass what we can.
        "process": original_ptr.process or "",
        "drawing": original_ptr.drawing or "",
        "dotr": original_ptr.dotr.strftime("%Y-%m-%d") if original_ptr.dotr else "",
        "ncsno": original_ptr.ncsno or "",
        "bqty": original_ptr.bqty or "",
        "mfg": original_ptr.mfg or "",
        "requestremarks": original_ptr.requestremarks or "",
        "mmodel": original_ptr.mmodel or "",
        "typemachine": original_ptr.typemachine or "",
        "req_path": original_ptr.req_path or "",
        
        # Revision specifics
        "current_rev_no": (rev.revisionNo or 0) + 1,
        "last_obsolete_date": rev.Vdate.strftime("%d-%m-%Y") if rev.Vdate else "",
    }
    
    return JsonResponse(data)


# ═══════════════════════════ Manager & Designer ═══════════════════════════


@login_required
@user_passes_test(is_manager)
@never_cache
def manager_approval_view(request):
    """Render the Manager Approval page with real PTR data (status=1 = pending)."""
    selected_unit = request.GET.get("unit", "")

    # Fetch pending requests (status=1)
    pending = PTRRequest.objects.filter(status=1)
    if selected_unit:
        pending = pending.filter(unit=selected_unit)

    from .models import TblRevision
    obsolete_qs = TblRevision.objects.filter(isapproved__isnull=True)
    if selected_unit:
        obsolete_qs = obsolete_qs.filter(ptr__unit=selected_unit)

    # Build enriched rows with department name + requestor email
    from django.contrib.auth.models import User
    
    ptr_rows = []
    tcmr_rows = []
    
    for ptr in pending:
        # Look up department name
        dept_name = ""
        try:
            dept = Department.objects.get(dID=ptr.deptid)
            dept_name = dept.DeptName
        except Department.DoesNotExist:
            dept_name = f"Dept #{ptr.deptid}"

        # Look up requestor email
        requestor_email = ""
        try:
            user = User.objects.get(username=ptr.createdby)
            requestor_email = user.email or ptr.createdby
        except User.DoesNotExist:
            requestor_email = ptr.createdby

        # Look up project name
        project_name = ""
        try:
            proj = Project.objects.get(PID=ptr.projectid)
            project_name = proj.prjname or ""
        except Project.DoesNotExist:
            project_name = f"Project #{ptr.projectid}"

        row_data = {
            "ptr": ptr,
            "request_type": "TCMR" if ptr.istcmr else "PTR",
            "dept_name": dept_name,
            "project_name": project_name,
            "requestor_email": requestor_email,
        }
        
        if ptr.istcmr:
            tcmr_rows.append(row_data)
        else:
            ptr_rows.append(row_data)

    obsolete_rows = []
    for rev in obsolete_qs:
        ptr = rev.ptr
        try:
            dept = Department.objects.get(dID=ptr.deptid)
            dept_name = dept.DeptName
        except Department.DoesNotExist:
            dept_name = f"Dept #{ptr.deptid}"
            
        obsolete_rows.append({
            "rev": rev,
            "ptr": ptr,
            "dept_name": dept_name,
        })

    return render(request, "ptr_app/manager_approval.html", {
        "ptr_rows": ptr_rows,
        "tcmr_rows": tcmr_rows,
        "obsolete_rows": obsolete_rows,
        "selected_unit": selected_unit,
    })


@login_required
@user_passes_test(is_manager)
@never_cache
def approve_request(request, pk):
    """Approve a PTR request — status 1 → 2."""
    if request.method == "POST":
        ptr = get_object_or_404(PTRRequest, pk=pk, status=1)
        ptr.status = 2
        ptr.aprrovedmanager = request.user.username
        from django.utils import timezone
        ptr.mappprovedate = timezone.now()
        ptr.save()
        messages.success(request, f"PTR-{ptr.pk:04d} approved successfully.")
    return redirect("manager_approval")


@login_required
@user_passes_test(is_manager)
@never_cache
def reject_request(request, pk):
    """Reject a PTR request — status 1 → -1."""
    if request.method == "POST":
        ptr = get_object_or_404(PTRRequest, pk=pk, status=1)
        ptr.status = -1
        ptr.statusremarks = request.POST.get("reason", "Rejected by manager")
        ptr.save()
        messages.success(request, f"PTR-{ptr.pk:04d} rejected.")
    return redirect("manager_approval")

@login_required
@user_passes_test(is_manager)
@never_cache
def approve_obsolete_view(request, pk):
    """Approve an obsolete fixture request, making it available for TCMR generation."""
    from .models import TblRevision
    from django.utils import timezone
    if request.method == "POST":
        rev = get_object_or_404(TblRevision, pk=pk)
        
        if rev.isapproved is None or rev.isapproved is False:
            rev.isapproved = True
            rev.Approver = request.user.username
            rev.Vdate = timezone.now().date()
            rev.save()
            messages.success(request, f"Obsolete Request for PTR-{rev.ptr.pk:04d} approved. (TCMR unlocked)")
    return redirect("manager_approval")


@login_required
@user_passes_test(is_manager)
@never_cache
def manager_edit_request(request, pk):
    """Allow manager to edit a pending PTR request before approval."""
    ptr = get_object_or_404(PTRRequest, pk=pk, status=1)

    if request.method == "POST":
        # Update editable fields from POST data
        ptr.process = request.POST.get("process", ptr.process)
        ptr.drawing = request.POST.get("drawing", ptr.drawing)
        ptr.ncsno = request.POST.get("ncsno", ptr.ncsno)
        ptr.bqty = request.POST.get("bqty", ptr.bqty)
        ptr.mfg = request.POST.get("mfg", ptr.mfg)
        ptr.requestremarks = request.POST.get("requestremarks", ptr.requestremarks)
        ptr.mmodel = request.POST.get("mmodel", ptr.mmodel)
        ptr.typemachine = request.POST.get("typemachine", ptr.typemachine)

        dotr_val = request.POST.get("dotr", "")
        if dotr_val:
            ptr.dotr = dotr_val

        ptr.save()
        messages.success(request, f"PTR-{ptr.pk:04d} updated successfully.")
        return redirect("manager_approval")

    # GET — render the edit form
    # Look up display names
    dept_name = ""
    try:
        dept = Department.objects.get(dID=ptr.deptid)
        dept_name = dept.DeptName
    except Department.DoesNotExist:
        dept_name = f"Dept #{ptr.deptid}"

    project_name = ""
    try:
        proj = Project.objects.get(PID=ptr.projectid)
        project_name = proj.prjname or ""
    except Project.DoesNotExist:
        project_name = ""

    customer_name = ""
    try:
        cust = Customer.objects.get(CustID=ptr.customerid)
        customer_name = cust.CustomerName
    except Customer.DoesNotExist:
        customer_name = ""

    # Get machine options for the unit
    machine_models = MachineModel.objects.filter(unit=ptr.unit)
    machine_types = TypeOfMachine.objects.filter(unit=ptr.unit)

    return render(request, "ptr_app/manager_edit_request.html", {
        "ptr": ptr,
        "dept_name": dept_name,
        "project_name": project_name,
        "customer_name": customer_name,
        "machine_models": machine_models,
        "machine_types": machine_types,
    })

@login_required
@user_passes_test(is_designer_or_admin)
@never_cache
def designer_entry_list_view(request):
    """Show PTRs assigned to designer (status=3 or 4) for the PTR Entry section."""
    from .models import PtrAdmin as PtrAdminModel

    selected_unit = request.GET.get("unit", "")

    # status=3: designer assigned, status=4: designer submitted (tooling exec pending)
    assigned = PTRRequest.objects.filter(status__in=[3, 4])
    if selected_unit:
        assigned = assigned.filter(unit=selected_unit)

    rows = []
    for idx, ptr in enumerate(assigned, 1):
        # Get PtrAdmin assignment details (planned dates, designer)
        admin_entry = PtrAdminModel.objects.filter(ptr=ptr).first()
        dept_name = ""
        try:
            dept = Department.objects.get(dID=ptr.deptid)
            dept_name = dept.DeptName
        except Department.DoesNotExist:
            dept_name = ""

        project_name = ""
        try:
            proj = Project.objects.get(PID=ptr.projectid)
            project_name = proj.prjname or ""
        except Project.DoesNotExist:
            project_name = ""

        rows.append({
            "sno": idx,
            "ptr": ptr,
            "dept_name": dept_name,
            "project_name": project_name,
            "admin": admin_entry,
        })

    return render(request, "ptr_app/designer_entry_list.html", {
        "rows": rows,
        "selected_unit": selected_unit,
    })


@login_required
@never_cache
def designer_entry_detail_view(request, req_id):
    """3-panel detail view — designer, tooling exec, tool room forms."""
    from .models import (
        PtrAdmin as PtrAdminModel, DesignerEntry, CriticalSpares
    )

    ptr = get_object_or_404(PTRRequest, pk=req_id, status__in=[3, 4, 5, 6, 7])
    admin_entry = PtrAdminModel.objects.filter(ptr=ptr).first()

    # Get or create the single DesignerEntry row for this PTR
    de, _ = DesignerEntry.objects.get_or_create(ptr=ptr)

    # Existing critical spares
    critical_spares = CriticalSpares.objects.filter(ptr=ptr)

    # Lookup display data
    dept_name, customer_name, project_name = "", "", ""
    try:
        dept_name = Department.objects.get(dID=ptr.deptid).DeptName
    except Department.DoesNotExist:
        pass
    try:
        customer_name = Customer.objects.get(CustID=ptr.customerid).CustomerName
    except Customer.DoesNotExist:
        pass
    try:
        project_name = Project.objects.get(PID=ptr.projectid).prjname or ""
    except Project.DoesNotExist:
        pass

    if request.method == "POST":
        form_type = request.POST.get("form_type", "")

        if form_type == "designer":
            # Form 1: Designer
            de.ASD = request.POST.get("ASD") or None
            de.ACD = request.POST.get("ACD") or None
            de.actualHr = int(request.POST.get("actualHr", 0) or 0)
            de.DescTooling = request.POST.get("DescTooling", "")
            de.TINo = request.POST.get("TINo", "")
            de.ECost = request.POST.get("ECost", "")
            de.Path = request.POST.get("Path", "")
            de.ToolMaterial = request.POST.get("ToolMaterial", "")
            de.toollife = request.POST.get("toollife", "")
            de.PMFrequency = int(request.POST.get("PMFrequency", 0) or 0)
            de.desremarks = request.POST.get("desremarks", "")
            de.isinhous = request.POST.get("isinhous") == "on"
            de.isVoiceSetup = request.POST.get("isVoiceSetup") == "on"
            de.chk_jaw = request.POST.get("chk_jaw") == "on"
            de.iscmc = request.POST.get("iscmc") == "on"
            de.status = True
            de.save()

            # Handle critical spares
            if de.iscmc:
                # Clear old spares and save new ones
                CriticalSpares.objects.filter(ptr=ptr).delete()
                drgnos = request.POST.getlist("spare_drgno")
                cmcodes = request.POST.getlist("spare_cmcode")
                descs = request.POST.getlist("spare_desc")
                qtys = request.POST.getlist("spare_qty")
                uoms = request.POST.getlist("spare_uom")
                for i in range(len(drgnos)):
                    if drgnos[i].strip() or (i < len(cmcodes) and cmcodes[i].strip()) or (i < len(descs) and descs[i].strip()):
                        CriticalSpares.objects.create(
                            ptr=ptr,
                            drgNO=drgnos[i].strip(),
                            cmcode=cmcodes[i].strip() if i < len(cmcodes) else "",
                            desc=descs[i].strip() if i < len(descs) else "",
                            qty=int(qtys[i]) if i < len(qtys) and qtys[i].strip() else None,
                            UOM=uoms[i].strip() if i < len(uoms) else "",
                        )

            if ptr.status < 4:
                ptr.status = 4  # Designer submitted → Tooling Exec pending
                ptr.save()
                messages.success(request, f"{ptr.ptrno or f'PTR-{ptr.pk:04d}'} — Designer entry submitted.")
            else:
                messages.success(request, f"{ptr.ptrno or f'PTR-{ptr.pk:04d}'} — Designer entry updated by Admin.")
            return redirect("designer_entry_list")

        elif form_type == "tooling_exec":
            # Form 2: Tooling Executive
            de.waitingroi = request.POST.get("waitingroi") == "on"
            de.roino = request.POST.get("roino", "")
            de.consumablecode = request.POST.get("consumablecode", "")
            de.mprDate = request.POST.get("mprDate") or None
            de.MPRNo = request.POST.get("MPRNo", "")
            de.suplier = request.POST.get("suplier", "")
            de.pocost = request.POST.get("pocost", "")
            de.eddate = request.POST.get("eddate") or None
            de.save()

            if ptr.status < 5:
                ptr.status = 5  # Tooling exec submitted
                ptr.save()
                messages.success(request, f"{ptr.ptrno or f'PTR-{ptr.pk:04d}'} — Tooling executive entry submitted.")
            else:
                messages.success(request, f"{ptr.ptrno or f'PTR-{ptr.pk:04d}'} — Tooling executive entry updated by Admin.")
            return redirect("designer_entry_list")

        elif form_type == "toolroom":
            # Form 3: Tool Room (later stage)
            de.Tfeedback = request.POST.get("Tfeedback", "")
            de.save()
            
            if ptr.status < 7:
                ptr.status = 7  # Feedback completed
                ptr.save()
                messages.success(request, f"{ptr.ptrno or f'PTR-{ptr.pk:04d}'} — Tool room feedback submitted.")
            else:
                messages.success(request, f"{ptr.ptrno or f'PTR-{ptr.pk:04d}'} — Tool room feedback updated by Admin.")
            return redirect("receiving_feedback")

        elif form_type == "common":
            # "Make Common with other TINo." submission
            de.bitcommon = True
            de.commonremarks = request.POST.get("commonremarks", "")
            common_id = request.POST.get("commonptrid")
            if common_id and common_id.isdigit():
                de.commonptrid = int(common_id)
            de.save()
            
            ptr.status = -3  # Rejected via Common
            ptr.save()
            messages.warning(request, f"{ptr.ptrno or f'PTR-{ptr.pk:04d}'} — Marked as Common with another TINO. Request closed.")
            return redirect("designer_entry_list")

        elif form_type == "shortclose":
            # Short close via Tool Room panel
            de.latereason = request.POST.get("shortclose_reason", "")
            de.save()

            ptr.status = -4  # Rejected via Short Close
            ptr.save()
            messages.warning(request, f"{ptr.ptrno or f'PTR-{ptr.pk:04d}'} — Request Short Closed.")
            return redirect("designer_entry_list")

    # Gather all TINOs for the "Make Common" modal dropdown
    all_tinos = DesignerEntry.objects.exclude(TINo__isnull=True).exclude(TINo="").values("ptr_id", "TINo")

    return render(request, "ptr_app/designer_entry_detail.html", {
        "ptr": ptr,
        "admin": admin_entry,
        "de": de,
        "critical_spares": critical_spares,
        "dept_name": dept_name,
        "customer_name": customer_name,
        "project_name": project_name,
        "all_tinos": all_tinos,
    })

# ═══════════════════════════ PTR Admin ═══════════════════════════


@login_required
@never_cache
def ptr_admin_list_view(request):
    """Show all manager-approved PTR requests (status=2) for admin assignment."""
    selected_unit = request.GET.get("unit", "")

    approved = PTRRequest.objects.filter(status=2)
    if selected_unit:
        approved = approved.filter(unit=selected_unit)

    from .models import PtrAdmin as PtrAdminModel
    # Exclude PTRs already assigned in PtrAdmin
    assigned_ids = PtrAdminModel.objects.values_list("ptr_id", flat=True)
    approved = approved.exclude(pk__in=assigned_ids)

    rows = []
    for idx, ptr in enumerate(approved, 1):
        dept_name = ""
        try:
            dept = Department.objects.get(dID=ptr.deptid)
            dept_name = dept.DeptName
        except Department.DoesNotExist:
            dept_name = f"Dept #{ptr.deptid}"

        customer_name = ""
        try:
            cust = Customer.objects.get(CustID=ptr.customerid)
            customer_name = cust.CustomerName
        except Customer.DoesNotExist:
            customer_name = ""

        project_name = ""
        try:
            proj = Project.objects.get(PID=ptr.projectid)
            project_name = proj.prjname or ""
        except Project.DoesNotExist:
            project_name = ""

        rows.append({
            "sno": idx,
            "ptr": ptr,
            "request_type": "TCMR" if ptr.istcmr else "PTR",
            "dept_name": dept_name,
            "customer_name": customer_name,
            "project_name": project_name,
        })

    return render(request, "ptr_app/ptr_admin_list.html", {
        "rows": rows,
        "selected_unit": selected_unit,
    })



@login_required
@never_cache
def ptr_admin_detail_view(request, pk):
    """Detail/assign page for a specific approved PTR request."""
    from .models import PtrAdmin as PtrAdminModel, EmployeeMapping

    ptr = get_object_or_404(PTRRequest, pk=pk, status=2)

    # Get designers from EmployeeMapping
    designers = EmployeeMapping.objects.filter(
        Role="Designer"
    ).values_list("EmpNumber", flat=True).distinct()

    # Generate PTR number for DISPLAY only — NOT saved yet
    # It will only be saved permanently when Assign is clicked
    ptr_number = _generate_ptr_number(ptr.unit or "ASMIPL")

    # Lookup display data
    dept_name = ""
    try:
        dept = Department.objects.get(dID=ptr.deptid)
        dept_name = dept.DeptName
    except Department.DoesNotExist:
        dept_name = ""

    customer_name = ""
    try:
        cust = Customer.objects.get(CustID=ptr.customerid)
        customer_name = cust.CustomerName
    except Customer.DoesNotExist:
        customer_name = ""

    project_name = ""
    try:
        proj = Project.objects.get(PID=ptr.projectid)
        project_name = proj.prjname or ""
    except Project.DoesNotExist:
        project_name = ""

    if request.method == "POST":
        action = request.POST.get("action", "")

        if action == "assign":
            # NOW lock in the PTR number permanently
            assigned_ptrno = request.POST.get("ptrno", ptr_number)
            ptr.ptrno = assigned_ptrno
            ptr.status = 3  # Assigned to designer
            ptr.save()

            PtrAdminModel.objects.create(
                ptrno=assigned_ptrno,
                desginer=request.POST.get("designer", ""),
                pldstartdate=request.POST.get("pldstartdate") or None,
                pldcompdate=request.POST.get("pldcompdate"),
                remarks=request.POST.get("remarks", ""),
                shortclose=False,
                status=True,
                ptr=ptr,
                plannedHour=int(request.POST.get("plannedHour", 0) or 0),
            )
            messages.success(request, f"{assigned_ptrno} assigned to designer successfully.")
            return redirect("ptr_admin_list")

        elif action == "reject":
            ptr.ptrno = None  # No PTR number for rejected
            ptr.status = -1
            ptr.statusremarks = "Rejected by PTR Admin"
            ptr.save()
            messages.success(request, f"PTR-{ptr.pk:04d} rejected.")
            return redirect("ptr_admin_list")

        elif action == "shortclose":
            ptr.ptrno = None  # No PTR number for short closed
            ptr.status = -2
            ptr.shortclose = True
            ptr.statusremarks = "Short closed by PTR Admin"
            ptr.save()
            messages.success(request, f"PTR-{ptr.pk:04d} short closed.")
            return redirect("ptr_admin_list")

    return render(request, "ptr_app/ptr_admin_detail.html", {
        "ptr": ptr,
        "ptr_number": ptr_number,
        "designers": designers,
        "dept_name": dept_name,
        "customer_name": customer_name,
        "project_name": project_name,
    })


# ═══════════════════════════ Receiving / PM / Tool Life ═══════════════════

def _generate_tvrn(unit):
    """Generate Tooling Verification Report Number: UNIT/FY/R01, R02, ...
    Same financial year logic as PTR number. Independent per unit."""
    from datetime import datetime
    from .models import ReceivingEntry
    import re

    now = datetime.now()
    if now.month >= 4:
        fy_start = now.year % 100
        fy_end = (now.year + 1) % 100
    else:
        fy_start = (now.year - 1) % 100
        fy_end = now.year % 100

    fy = f"{fy_start:02d}-{fy_end:02d}"
    prefix = f"{unit}/{fy}/R"

    max_serial = 0
    for entry in ReceivingEntry.objects.filter(tvrn__startswith=prefix):
        match = re.search(r'/R(\d+)$', entry.tvrn)
        if match:
            serial = int(match.group(1))
            if serial > max_serial:
                max_serial = serial

    next_serial = max_serial + 1
    return f"{prefix}{next_serial:02d}"


@login_required
@never_cache
def receiving_entry_view(request):
    """Show PTRs where tooling exec is done (status=5) for receiving."""
    from .models import (
        PtrAdmin as PtrAdminModel, DesignerEntry, ReceivingEntry
    )

    selected_unit = request.GET.get("unit", "")

    # status=5: tooling exec done → ready for receiving
    pending = PTRRequest.objects.filter(status=5)
    if selected_unit:
        pending = pending.filter(unit=selected_unit)

    rows = []
    for idx, ptr in enumerate(pending, 1):
        admin_entry = PtrAdminModel.objects.filter(ptr=ptr).first()
        de = DesignerEntry.objects.filter(ptr=ptr).first()
        dept_name = ""
        try:
            dept_name = Department.objects.get(dID=ptr.deptid).DeptName
        except Department.DoesNotExist:
            pass
        project_name = ""
        try:
            project_name = Project.objects.get(PID=ptr.projectid).prjname or ""
        except Project.DoesNotExist:
            pass

        # Check if already received
        received = ReceivingEntry.objects.filter(ptr=ptr).exists()

        rows.append({
            "sno": idx,
            "ptr": ptr,
            "dept_name": dept_name,
            "project_name": project_name,
            "admin": admin_entry,
            "de": de,
            "received": received,
        })

    return render(request, "ptr_app/receiving_entry.html", {
        "rows": rows,
        "selected_unit": selected_unit,
    })


@login_required
@never_cache
def receiving_entry_detail_view(request, req_id):
    """Receiving form — auto-fills PTR data, 4 manual fields + auto TVRN."""
    from .models import (
        PtrAdmin as PtrAdminModel, DesignerEntry, ReceivingEntry
    )

    ptr = get_object_or_404(PTRRequest, pk=req_id, status__in=[5, 6])
    admin_entry = PtrAdminModel.objects.filter(ptr=ptr).first()
    de = DesignerEntry.objects.filter(ptr=ptr).first()
    existing_recv = ReceivingEntry.objects.filter(ptr=ptr).first()

    # Auto-generate TVRN (only if not yet received)
    tvrn = existing_recv.tvrn if existing_recv else _generate_tvrn(ptr.unit or "ASMIPL")

    # Lookup display data
    dept_name, customer_name, project_name = "", "", ""
    try:
        dept_name = Department.objects.get(dID=ptr.deptid).DeptName
    except Department.DoesNotExist:
        pass
    try:
        customer_name = Customer.objects.get(CustID=ptr.customerid).CustomerName
    except Customer.DoesNotExist:
        pass
    try:
        project_name = Project.objects.get(PID=ptr.projectid).prjname or ""
    except Project.DoesNotExist:
        pass

    if request.method == "POST" and not existing_recv:
        from django.utils import timezone

        supplier = request.POST.get("supplier", "")
        grnno = request.POST.get("grnno", "")
        grndate = request.POST.get("grndate") or None
        rcvdate = request.POST.get("rcvdate") or None
        inspected_by = request.POST.get("inspected_by", "")

        ReceivingEntry.objects.create(
            ptr=ptr,
            rcvdate=rcvdate if rcvdate else timezone.now(),
            tvrn=tvrn,
            grnno=grnno,
            grndate=grndate,
            ModifiedBy=inspected_by,
            ModifiedDate=timezone.now(),
        )

        # Update supplier in DesignerEntry if provided
        if de and supplier:
            de.suplier = supplier
            de.save()

        ptr.status = 6  # Received
        ptr.save()
        messages.success(request, f"{ptr.ptrno} — Receiving entry submitted. TVRN: {tvrn}")
        return redirect("receiving_entry")

    return render(request, "ptr_app/receiving_entry_detail.html", {
        "ptr": ptr,
        "admin": admin_entry,
        "de": de,
        "existing_recv": existing_recv,
        "tvrn": tvrn,
        "dept_name": dept_name,
        "customer_name": customer_name,
        "project_name": project_name,
    })


@login_required
@never_cache
def receiving_feedback_view(request):
    """Show PTRs where receiving is done (status=6) for Tool Room feedback.
    Also shows status 7 for reference."""
    from .models import (
        PtrAdmin as PtrAdminModel, DesignerEntry, ReceivingEntry
    )

    selected_unit = request.GET.get("unit", "")
    comp_unit = request.GET.get("comp_unit", "")

    # status=6: received → ready for feedback
    pending = PTRRequest.objects.filter(status=6)
    if selected_unit:
        pending = pending.filter(unit=selected_unit)

    # status=7: feedback done
    completed = PTRRequest.objects.filter(status=7)
    if comp_unit:
        completed = completed.filter(unit=comp_unit)

    pending_rows = []
    completed_rows = []
    
    # Process Pending PTRs
    for idx, ptr in enumerate(pending, 1):
        admin_entry = PtrAdminModel.objects.filter(ptr=ptr).first()
        de = DesignerEntry.objects.filter(ptr=ptr).first()
        dept_name = ""
        try:
            dept_name = Department.objects.get(dID=ptr.deptid).DeptName
        except Department.DoesNotExist:
            pass
        project_name = ""
        try:
            project_name = Project.objects.get(PID=ptr.projectid).prjname or ""
        except Project.DoesNotExist:
            pass
        
        pending_rows.append({
            "sno": idx,
            "ptr": ptr,
            "dept_name": dept_name,
            "project_name": project_name,
            "admin": admin_entry,
            "de": de,
            "feedback_done": False,
        })

    # Process Completed PTRs
    for idx, ptr in enumerate(completed, 1):
        admin_entry = PtrAdminModel.objects.filter(ptr=ptr).first()
        de = DesignerEntry.objects.filter(ptr=ptr).first()
        dept_name = ""
        try:
            dept_name = Department.objects.get(dID=ptr.deptid).DeptName
        except Department.DoesNotExist:
            pass
        project_name = ""
        try:
            project_name = Project.objects.get(PID=ptr.projectid).prjname or ""
        except Project.DoesNotExist:
            pass
        
        completed_rows.append({
            "sno": idx,
            "ptr": ptr,
            "dept_name": dept_name,
            "project_name": project_name,
            "admin": admin_entry,
            "de": de,
            "feedback_done": True,
        })

    return render(request, "ptr_app/receiving_feedback.html", {
        "pending_rows": pending_rows,
        "completed_rows": completed_rows,
        "selected_unit": selected_unit,
        "comp_unit": comp_unit,
    })

# ═══════════════════════════ Short Close ═══════════════════════════


@login_required
@never_cache
def short_close_view(request):
    """Short Close: mark any active PTR (status >= 3) as short closed (status -2)."""
    if request.method == "POST":
        ptr_id = request.POST.get("ptr_id")
        if not ptr_id:
            messages.error(request, "No PTR selected.")
            return redirect("short_close")

        ptr = get_object_or_404(PTRRequest, pk=int(ptr_id))
        if ptr.status < 3:
            messages.error(request, "This PTR is not eligible for Short Close.")
            return redirect("short_close")

        ptr.status = -2
        ptr.shortclose = True
        ptr.statusremarks = "Short Closed"
        ptr.save()
        messages.success(request, f"{ptr.ptrno or f'PTR-{ptr.pk:04d}'} — Short Closed successfully.")
        return redirect("short_close")

    return render(request, "ptr_app/short_close.html")


@login_required
def api_active_ptrs_by_unit(request):
    """GET /api/active-ptrs/?unit=X → JSON list of active PTRs (status >= 3) for the unit."""
    unit = request.GET.get("unit", "")
    qs = PTRRequest.objects.filter(status__gte=3)
    if unit:
        qs = qs.filter(unit=unit)
    ptrs = []
    for ptr in qs.order_by("ptrno"):
        ptrs.append({
            "id": ptr.pk,
            "ptrno": ptr.ptrno or f"PTR-{ptr.pk:04d}",
        })
    return JsonResponse({"ptrs": ptrs})


def _generate_breakdown_number(unit):
    import datetime
    import re
    from .models import Breakdown
    
    now = datetime.datetime.now()
    if now.month >= 4:
        fy_start = now.year % 100
        fy_end = (now.year + 1) % 100
    else:
        fy_start = (now.year - 1) % 100
        fy_end = now.year % 100
    
    fy = f"{fy_start:02d}-{fy_end:02d}"
    prefix = f"{unit}/{fy}/BR"

    max_serial = 0
    existing_entries = Breakdown.objects.filter(reportNo__startswith=prefix)
    
    for entry in existing_entries:
        match = re.search(r'BR(\d+)$', entry.reportNo or "")
        if match:
            serial = int(match.group(1))
            if serial > max_serial:
                max_serial = serial

    next_serial = max_serial + 1
    return f"{prefix}{next_serial:02d}"


@login_required
@never_cache
def breakdown_maintenance_view(request):
    """Render the Breakdown Maintenance Master-Detail view."""
    from .models import PTRRequest, DesignerEntry, Breakdown, EmployeeMapping
    from django.contrib import messages
    from django.shortcuts import redirect
    
    selected_unit = request.GET.get("unit", "")
    
    # 1. Provide TINo list: status=7, NOT obsolete
    valid_ptrs = PTRRequest.objects.filter(status=7).exclude(isobsolete=True)
    if selected_unit:
        valid_ptrs = valid_ptrs.filter(unit=selected_unit)
        
    tino_list = []
    for ptr in valid_ptrs:
        de = DesignerEntry.objects.filter(ptr=ptr).first()
        if de and de.TINo:
            tino_list.append({"tino": de.TINo, "ptr_id": ptr.pk, "unit": ptr.unit or "ASMIPL"})
            
    # 2. Provide Designers for Approval dropdown
    designers = EmployeeMapping.objects.filter(Role="Designer")
    
    # 3. Handle POST
    if request.method == "POST":
        ptr_id = request.POST.get("ptr_id")
        reason = request.POST.get("reason", "")
        bd_date_str = request.POST.get("bd_date")
        downtime_val = request.POST.get("downtime") or 0.0
        status_val = request.POST.get("status", "")
        approver = request.POST.get("approver", "")
        
        if ptr_id:
            try:
                target_ptr = PTRRequest.objects.get(pk=ptr_id)
                target_de = DesignerEntry.objects.filter(ptr=target_ptr).first()
                tino_val = target_de.TINo if target_de else ""
                
                # Compute Auto-Gen ID based on PTR's Unit (or strictly ASMIPL if none)
                new_report_no = _generate_breakdown_number(target_ptr.unit or "ASMIPL")
                
                Breakdown.objects.create(
                    ptr=target_ptr,
                    Tino=tino_val,
                    bd_date=bd_date_str if bd_date_str else None,
                    reason=reason,
                    reportNo=new_report_no,
                    Status=status_val,
                    downtime=downtime_val,
                    Approved=approver
                )
                messages.success(request, f"Successfully created breakdown entry {new_report_no} for TINo {tino_val}.")
                return redirect(f"/preventive-maintenance/breakdown/?unit={selected_unit}")
            except Exception as e:
                messages.error(request, f"Error saving Breakdown entry: {e}")

    # 4. Filter breakdown existing records for table
    rows = Breakdown.objects.all().order_by('-pk')
    if selected_unit:
        rows = rows.filter(ptr__unit=selected_unit)
        
    # 5. Pre-compute the next report number for visual display dynamically targeting all supported units
    next_report_no = {
        "ASMIPL": _generate_breakdown_number("ASMIPL"),
        "ASYM": _generate_breakdown_number("ASYM"),
    }
        
    return render(request, "ptr_app/breakdown_maintenance.html", {
        "tino_list": tino_list,
        "designers": designers,
        "rows": rows,
        "selected_unit": selected_unit,
        "next_report_no": next_report_no
    })


@login_required
@never_cache
def pm_due_view(request):
    """Monthly PM Due Count for next 36 months, split by unit."""
    import datetime
    import calendar
    import re
    from collections import OrderedDict
    from .models import PTRRequest, DesignerEntry, ReceivingEntry, PreventM

    def add_months(sourcedate, months):
        month = sourcedate.month - 1 + months
        year = int(sourcedate.year + month / 12)
        month = month % 12 + 1
        day = min(sourcedate.day, calendar.monthrange(year, month)[1])
        return datetime.date(year, month, day)

    today = datetime.date.today()

    # Build 36-month bucket keys starting from the current month
    month_names = [
        'January', 'February', 'March', 'April', 'May', 'June',
        'July', 'August', 'September', 'October', 'November', 'December'
    ]
    buckets = OrderedDict()
    for i in range(36):
        m = today.month - 1 + i
        year = today.year + m // 12
        month = m % 12 + 1
        key = (year, month)
        buckets[key] = {'asmipl': 0, 'asym': 0}

    # Query all Status-7 PTRs (completed / active tools)
    base_qs = PTRRequest.objects.filter(status=7).order_by('-pk')

    for ptr in base_qs:
        de = DesignerEntry.objects.filter(ptr=ptr).first()
        rcv = ReceivingEntry.objects.filter(ptr=ptr).first()

        if not de or not rcv or not rcv.rcvdate:
            continue

        freq = de.PMFrequency or 0
        if freq <= 0:
            continue

        tl_str = de.toollife or ""
        m_match = re.search(r'\d+', str(tl_str))
        toollife_years = int(m_match.group()) if m_match else 0
        if toollife_years <= 0:
            continue

        rcv_date_only = rcv.rcvdate.date() if isinstance(rcv.rcvdate, datetime.datetime) else rcv.rcvdate

        # Skip expired tools
        tool_life_expires = add_months(rcv_date_only, toollife_years * 12)
        if tool_life_expires <= today:
            continue

        max_inspections = toollife_years * freq
        interval_months = 12 // freq

        # Pre-fetch existing PM dates for this tool
        existing_pm_dates = set(PreventM.objects.filter(ptr=ptr).values_list('pmdate', flat=True))

        unit_key = 'asmipl' if ptr.unit == 'ASMIPL' else 'asym'

        for i in range(1, max_inspections + 1):
            target_date = add_months(rcv_date_only, interval_months * i)
            # Only count PM dues that haven't been completed yet
            if target_date in existing_pm_dates:
                continue
            bucket_key = (target_date.year, target_date.month)
            if bucket_key in buckets:
                buckets[bucket_key][unit_key] += 1

    # Build rows for template
    pm_due_rows = []
    total_asmipl = 0
    total_asym = 0
    total_all = 0

    for (year, month), counts in buckets.items():
        asmipl_c = counts['asmipl']
        asym_c = counts['asym']
        row_total = asmipl_c + asym_c
        total_asmipl += asmipl_c
        total_asym += asym_c
        total_all += row_total
        pm_due_rows.append({
            'month_name': month_names[month - 1],
            'asmipl': asmipl_c,
            'asym': asym_c,
            'total': row_total,
            'year_month': f"{year}-{month:02d}",
        })

    return render(request, "ptr_app/pm_due.html", {
        "pm_due_rows": pm_due_rows,
        "total_asmipl": total_asmipl,
        "total_asym": total_asym,
        "total_all": total_all,
    })


@login_required
@never_cache
def export_pm_due_excel(request):
    """Export PM Due 36-month forecast as a styled Excel file."""
    import datetime
    import calendar
    import re
    from collections import OrderedDict
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from .models import PTRRequest, DesignerEntry, ReceivingEntry, PreventM

    def add_months(sourcedate, months):
        month = sourcedate.month - 1 + months
        year = int(sourcedate.year + month / 12)
        month = month % 12 + 1
        day = min(sourcedate.day, calendar.monthrange(year, month)[1])
        return datetime.date(year, month, day)

    today = datetime.date.today()
    month_names = [
        'January', 'February', 'March', 'April', 'May', 'June',
        'July', 'August', 'September', 'October', 'November', 'December'
    ]

    buckets = OrderedDict()
    for i in range(36):
        m = today.month - 1 + i
        year = today.year + m // 12
        month = m % 12 + 1
        buckets[(year, month)] = {'asmipl': 0, 'asym': 0}

    base_qs = PTRRequest.objects.filter(status=7).order_by('-pk')
    for ptr in base_qs:
        de = DesignerEntry.objects.filter(ptr=ptr).first()
        rcv = ReceivingEntry.objects.filter(ptr=ptr).first()
        if not de or not rcv or not rcv.rcvdate:
            continue
        freq = de.PMFrequency or 0
        if freq <= 0:
            continue
        tl_str = de.toollife or ""
        m_match = re.search(r'\d+', str(tl_str))
        toollife_years = int(m_match.group()) if m_match else 0
        if toollife_years <= 0:
            continue
        rcv_date_only = rcv.rcvdate.date() if isinstance(rcv.rcvdate, datetime.datetime) else rcv.rcvdate
        tool_life_expires = add_months(rcv_date_only, toollife_years * 12)
        if tool_life_expires <= today:
            continue
        max_inspections = toollife_years * freq
        interval_months = 12 // freq
        existing_pm_dates = set(PreventM.objects.filter(ptr=ptr).values_list('pmdate', flat=True))
        unit_key = 'asmipl' if ptr.unit == 'ASMIPL' else 'asym'
        for i in range(1, max_inspections + 1):
            target_date = add_months(rcv_date_only, interval_months * i)
            if target_date in existing_pm_dates:
                continue
            bucket_key = (target_date.year, target_date.month)
            if bucket_key in buckets:
                buckets[bucket_key][unit_key] += 1

    # ── Build Excel Workbook ──
    wb = Workbook()
    ws = wb.active
    ws.title = "PM Due Forecast"

    # Styling
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    title_font = Font(name='Calibri', bold=True, size=14, color='1F4E79')
    data_font = Font(name='Calibri', size=11)
    total_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    total_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )
    alt_fill = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')

    # Column widths
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 14

    # Title row
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = f"Monthly PM Due Count for Next 36 Months (Generated: {today.strftime('%d-%b-%Y')})"
    title_cell.font = title_font
    title_cell.alignment = center

    # Header row
    headers = ['Month', 'ASMIPL PM Due', 'ASYM PM Due', 'Total PM Due', 'Year']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # Data rows
    total_asmipl = 0
    total_asym = 0
    total_all = 0
    row_num = 4

    for idx, ((year, month), counts) in enumerate(buckets.items()):
        asmipl_c = counts['asmipl']
        asym_c = counts['asym']
        row_total = asmipl_c + asym_c
        total_asmipl += asmipl_c
        total_asym += asym_c
        total_all += row_total

        row_fill = alt_fill if idx % 2 == 1 else None

        vals = [month_names[month - 1], asmipl_c, asym_c, row_total, f"{year}-{month:02d}"]
        aligns = [center, center, center, center, center]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = data_font
            cell.alignment = aligns[col - 1]
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill
        row_num += 1

    # Grand Total row
    total_vals = ['Grand Total', total_asmipl, total_asym, total_all, '']
    for col, val in enumerate(total_vals, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = center
        cell.border = thin_border

    # Serve response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"PM_Due_Forecast_{today.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
@never_cache
def pm_plan_view(request):
    """Render the PM Plan Master-Detail view."""
    import datetime
    import calendar
    import re
    from .models import PTRRequest, DesignerEntry, ReceivingEntry, PreventM

    def add_months(sourcedate, months):
        month = sourcedate.month - 1 + months
        year = int(sourcedate.year + month / 12)
        month = month % 12 + 1
        day = min(sourcedate.day, calendar.monthrange(year,month)[1])
        return datetime.date(year, month, day)

    timing = request.GET.get('timing', 'current')
    selected_unit = request.GET.get('unit', '')

    base_qs = PTRRequest.objects.filter(status=7).order_by('-pk')
    if selected_unit:
        base_qs = base_qs.filter(unit=selected_unit)

    today = datetime.date.today()
    
    if today.month == 12:
        next_month_start = datetime.date(today.year + 1, 1, 1)
        next_month_end = datetime.date(today.year + 1, 1, 31)
    else:
        next_month_start = datetime.date(today.year, today.month + 1, 1)
        next_month_end = datetime.date(today.year, today.month + 1, calendar.monthrange(today.year, today.month + 1)[1])
        
    if today.month == 1:
        last_month_num = 12
        last_month_year = today.year - 1
    else:
        last_month_num = today.month - 1
        last_month_year = today.year
    
    pm_rows = []
    
    for ptr in base_qs:
        de = DesignerEntry.objects.filter(ptr=ptr).first()
        rcv = ReceivingEntry.objects.filter(ptr=ptr).first()
        
        if not de or not rcv or not rcv.rcvdate:
            continue
            
        freq = de.PMFrequency or 0
        if freq <= 0: continue
        
        tl_str = de.toollife or ""
        m = re.search(r'\d+', str(tl_str))
        toollife_years = int(m.group()) if m else 0
        
        if toollife_years <= 0: continue
        
        rcv_date_only = rcv.rcvdate.date() if isinstance(rcv.rcvdate, datetime.datetime) else rcv.rcvdate
        
        # Tools whose tool life is expired must not be displayed in PM plan
        tool_life_expires = add_months(rcv_date_only, toollife_years * 12)
        if tool_life_expires <= today:
            continue
            
        max_inspections = toollife_years * freq
        interval_months = 12 // freq
        
        upcoming_date = None
        include_row = False
        # Pre-fetch existing inspection schedule natively avoiding nested database sweeps
        existing_pm_dates = set(PreventM.objects.filter(ptr=ptr).values_list('pmdate', flat=True))
        
        for i in range(1, max_inspections + 1):
            target_date = add_months(rcv_date_only, interval_months * i)
            # Local RAM lookup instead of deep network querying
            if target_date not in existing_pm_dates:
                is_match = False
                if timing == 'past':
                    if target_date.year == last_month_year and target_date.month == last_month_num:
                        is_match = True
                elif timing == 'current':
                    if target_date.year == today.year and target_date.month == today.month:
                        is_match = True
                elif timing == 'next':
                    if target_date >= next_month_start and target_date <= next_month_end:
                        is_match = True
                        
                if is_match:
                    upcoming_date = target_date
                    include_row = True
                    break
                    
        if include_row and upcoming_date:
            past_pm = PreventM.objects.filter(ptr=ptr).order_by('-pmdate').first()
            last_report = past_pm.pmreportno if past_pm else "—"

            pm_rows.append({
                'ptr': ptr,
                'rcvdate': rcv_date_only,
                'plan_date': upcoming_date,
                'freq': freq,
                'tino': de.TINo,
                'tfeedback': de.Tfeedback,
                'last_report': last_report
            })
            
    return render(request, "ptr_app/pm_plan.html", {
        "pm_rows": pm_rows,
        "timing": timing,
        "selected_unit": selected_unit
    })

@login_required
@never_cache
def pm_plan_detail_view(request, pk):
    """Render the 3 forms for PM Inspection."""
    import datetime
    import calendar
    import re
    from .models import PTRRequest, DesignerEntry, ReceivingEntry, PreventM, EmployeeMapping

    ptr = get_object_or_404(PTRRequest, pk=pk)
    de = DesignerEntry.objects.filter(ptr=ptr).first()
    rcv = ReceivingEntry.objects.filter(ptr=ptr).first()
    designers = EmployeeMapping.objects.filter(Role="Designer")

    def add_months(sourcedate, months):
        month = sourcedate.month - 1 + months
        year = int(sourcedate.year + month / 12)
        month = month % 12 + 1
        day = min(sourcedate.day, calendar.monthrange(year,month)[1])
        return datetime.date(year, month, day)

    # Re-calculate dates
    rcvdate_display = None
    tino = ""
    frequency = ""
    tfeedback = ""
    upcoming_date = None
    tool_life_expires = None
    
    if de and rcv and rcv.rcvdate:
        tino = de.TINo
        frequency = de.PMFrequency or 0
        tfeedback = de.Tfeedback
        rcvdate_display = rcv.rcvdate.date() if isinstance(rcv.rcvdate, datetime.datetime) else rcv.rcvdate
        
        tl_str = de.toollife or ""
        m = re.search(r'\d+', str(tl_str))
        toollife_years = int(m.group()) if m else 0
        
        if toollife_years > 0 and frequency > 0:
            tool_life_expires = add_months(rcvdate_display, toollife_years * 12)
            max_inspections = toollife_years * frequency
            interval_months = 12 // frequency
            
            # Check if a specific target date was passed from the list filter context
            target_str = request.GET.get("target")
            if target_str:
                try:
                    upcoming_date = datetime.datetime.strptime(target_str, "%Y-%m-%d").date()
                except:
                    pass
                    
            # If no target passed statically, fallback to mathematically absolute closest target
            if not upcoming_date:
                for i in range(1, max_inspections + 1):
                    target_date = add_months(rcvdate_display, interval_months * i)
                    if not PreventM.objects.filter(ptr=ptr, pmdate=target_date).exists() and not PreventM.objects.filter(ptr=ptr, actualdate__isnull=False, pmdate=target_date).exists():
                        upcoming_date = target_date
                        break
                        
    if request.method == "POST":
        pm_ins_date = request.POST.get("pm_ins_date")
        pmreportno = request.POST.get("pmreportno", "")
        rmk = request.POST.get("rmk", "")
        status = request.POST.get("status", "")
        approver = request.POST.get("approver", "")
        inspector = request.POST.get("inspector", "")
        
        if pm_ins_date and upcoming_date:
            PreventM.objects.create(
                pmdate=upcoming_date, 
                pmreportno=pmreportno,
                ptr=ptr,
                actualdate=pm_ins_date,
                rmk=rmk,
                inspector=inspector,
                Approver=approver,
                status=status,
                TINO=tino
            )
            messages.success(request, f"Successfully submitted actual PM inspection for PTR-{ptr.ptrno or f'{ptr.pk:04d}'}.")
            return redirect("pm_plan")

    return render(request, "ptr_app/pm_plan_detail.html", {
        "ptr": ptr,
        "rcvdate": rcvdate_display,
        "tino": tino,
        "frequency": frequency,
        "tfeedback": tfeedback,
        "upcoming_date": upcoming_date,
        "tool_life_expires": tool_life_expires,
        "tool_life_duration": de.toollife if de else "",
        "designers": designers,
    })
@login_required
@never_cache
def tool_life_extension_view(request):
    """Render the Tool Life Extension Master-Detail view."""
    import datetime
    import calendar
    import re
    from .models import PTRRequest, DesignerEntry, ReceivingEntry

    def add_months(sourcedate, months):
        month = sourcedate.month - 1 + months
        year = int(sourcedate.year + month / 12)
        month = month % 12 + 1
        day = min(sourcedate.day, calendar.monthrange(year,month)[1])
        return datetime.date(year, month, day)

    selected_unit = request.GET.get('unit', '')
    base_qs = PTRRequest.objects.filter(status=7).exclude(isobsolete=True).order_by('-pk')
    if selected_unit:
        base_qs = base_qs.filter(unit=selected_unit)

    today = datetime.date.today()
    expired_rows = []

    for ptr in base_qs:
        de = DesignerEntry.objects.filter(ptr=ptr).first()
        rcv = ReceivingEntry.objects.filter(ptr=ptr).first()
        
        if not de or not rcv or not rcv.rcvdate:
            continue
            
        tl_str = de.toollife or ""
        m = re.search(r'\d+', str(tl_str))
        toollife_years = int(m.group()) if m else 0
        
        if toollife_years <= 0:
            continue
            
        rcv_date_only = rcv.rcvdate.date() if isinstance(rcv.rcvdate, datetime.datetime) else rcv.rcvdate
        
        tool_life_expires = add_months(rcv_date_only, toollife_years * 12)
        
        if tool_life_expires <= today:
            expired_rows.append({
                'ptr': ptr,
                'tino': de.TINo,
                'material': de.ToolMaterial,
                'expiry_date': tool_life_expires,
                'rcvdate': rcv_date_only,
                'freq': de.PMFrequency,
                'toollife': de.toollife,
            })

    return render(request, "ptr_app/tool_life_extension.html", {
        "expired_rows": expired_rows,
        "selected_unit": selected_unit
    })

@login_required
@never_cache
def tool_life_extend_detail_view(request, pk):
    """Render the Tool Life Extension 2-form detail view WITHOUT active POST saving yet."""
    import datetime
    import calendar
    import re
    from .models import PTRRequest, DesignerEntry, ReceivingEntry, EmployeeMapping
    
    ptr = get_object_or_404(PTRRequest, pk=pk)
    de = DesignerEntry.objects.filter(ptr=ptr).first()
    rcv = ReceivingEntry.objects.filter(ptr=ptr).first()
    designers = EmployeeMapping.objects.filter(Role="Designer")
    
    def add_months(sourcedate, months):
        month = sourcedate.month - 1 + months
        year = int(sourcedate.year + month / 12)
        month = month % 12 + 1
        day = min(sourcedate.day, calendar.monthrange(year,month)[1])
        return datetime.date(year, month, day)
        
    tino = ""
    material = ""
    rcvdate_display = None
    expiry_date = None
    
    if de and rcv and rcv.rcvdate:
        tino = de.TINo
        material = de.ToolMaterial
        rcvdate_display = rcv.rcvdate.date() if isinstance(rcv.rcvdate, datetime.datetime) else rcv.rcvdate
        
        tl_str = de.toollife or ""
        m = re.search(r'\d+', str(tl_str))
        toollife_years = int(m.group()) if m else 0
        if toollife_years > 0:
            expiry_date = add_months(rcvdate_display, toollife_years * 12)
            
    return render(request, "ptr_app/tool_life_extend_detail.html", {
        "ptr": ptr,
        "tino": tino,
        "material": material,
        "rcvdate_display": rcvdate_display,
        "expiry_date": expiry_date,
        "designers": designers
    })


@login_required
@never_cache
def history_card_view(request, pk):
    """Render the native printable PDF layout for History Card."""
    from .models import PTRRequest, DesignerEntry, ReceivingEntry, PreventM, Breakdown, ToolLifeExtend
    ptr = get_object_or_404(PTRRequest, pk=pk)
    de = DesignerEntry.objects.filter(ptr=ptr).first()
    rcv = ReceivingEntry.objects.filter(ptr=ptr).first()
    
    tino = de.TINo if de else ""
    pm_frequency = de.PMFrequency if de else ""
    tool_life = de.toollife if de else ""
    part_name = ptr.partname
    part_number = ptr.partcode
    process = ptr.process
    tool_description = de.DescTooling if de else ""
    make_supplier = de.suplier if de else ""
    date_of_commission = rcv.rcvdate.date() if rcv and rcv.rcvdate else ""
    refer_report_no = rcv.tvrn if rcv else ""
    
    pm_events = PreventM.objects.filter(ptr=ptr).order_by('pmdate')
    bd_events = Breakdown.objects.filter(ptr=ptr).order_by('bd_date')
    mod_events = ToolLifeExtend.objects.filter(PTRId=ptr).order_by('ExtendedOn')

    return render(request, "ptr_app/history_card_pdf.html", {
        "ptr": ptr,
        "tino": tino,
        "pm_frequency": pm_frequency,
        "tool_life": tool_life,
        "part_name": part_name,
        "tool_description": tool_description,
        "part_number": part_number,
        "process": process,
        "make_supplier": make_supplier,
        "date_of_commission": date_of_commission,
        "refer_report_no": refer_report_no,
        "pm_events": pm_events,
        "bd_events": bd_events,
        "mod_events": mod_events,
    })

# ═══════════════════════════ Dashboard ═══════════════════════════

@login_required
@never_cache
def dashboard_view(request):
    """Render the main Real-Time Dashboard UI."""
    import datetime
    import calendar
    import re
    from collections import OrderedDict
    from django.db.models import Count, Q
    from .models import PtrAdmin, PTRRequest, DesignerEntry, ReceivingEntry, PreventM

    # ── Designer Workload (Status 3 to 6) ──
    designer_workload = PtrAdmin.objects.filter(
        ptr__status__in=[3, 4, 5, 6]
    ).exclude(
        desginer__exact=''
    ).exclude(
        desginer__isnull=True
    ).values(
        'desginer'
    ).annotate(
        asmipl_count=Count('ptr', filter=Q(ptr__unit='ASMIPL')),
        asym_count=Count('ptr', filter=Q(ptr__unit='ASYM')),
        total_count=Count('ptr')
    ).order_by('desginer')

    total_asmipl = sum(d['asmipl_count'] for d in designer_workload)
    total_asym = sum(d['asym_count'] for d in designer_workload)
    grand_total = sum(d['total_count'] for d in designer_workload)

    # ── PM Due Forecast (next 36 months) ──
    def add_months(sourcedate, months):
        month = sourcedate.month - 1 + months
        year = int(sourcedate.year + month / 12)
        month = month % 12 + 1
        day = min(sourcedate.day, calendar.monthrange(year, month)[1])
        return datetime.date(year, month, day)

    today = datetime.date.today()
    month_names = [
        'January', 'February', 'March', 'April', 'May', 'June',
        'July', 'August', 'September', 'October', 'November', 'December'
    ]

    buckets = OrderedDict()
    for i in range(36):
        m = today.month - 1 + i
        year = today.year + m // 12
        month = m % 12 + 1
        buckets[(year, month)] = {'asmipl': 0, 'asym': 0}

    base_qs = PTRRequest.objects.filter(status=7).order_by('-pk')
    for ptr in base_qs:
        de = DesignerEntry.objects.filter(ptr=ptr).first()
        rcv = ReceivingEntry.objects.filter(ptr=ptr).first()
        if not de or not rcv or not rcv.rcvdate:
            continue
        freq = de.PMFrequency or 0
        if freq <= 0:
            continue
        tl_str = de.toollife or ""
        m_match = re.search(r'\d+', str(tl_str))
        toollife_years = int(m_match.group()) if m_match else 0
        if toollife_years <= 0:
            continue
        rcv_date_only = rcv.rcvdate.date() if isinstance(rcv.rcvdate, datetime.datetime) else rcv.rcvdate
        tool_life_expires = add_months(rcv_date_only, toollife_years * 12)
        if tool_life_expires <= today:
            continue
        max_inspections = toollife_years * freq
        interval_months = 12 // freq
        existing_pm_dates = set(PreventM.objects.filter(ptr=ptr).values_list('pmdate', flat=True))
        unit_key = 'asmipl' if ptr.unit == 'ASMIPL' else 'asym'
        for i in range(1, max_inspections + 1):
            target_date = add_months(rcv_date_only, interval_months * i)
            if target_date in existing_pm_dates:
                continue
            bucket_key = (target_date.year, target_date.month)
            if bucket_key in buckets:
                buckets[bucket_key][unit_key] += 1

    pm_due_rows = []
    pm_total_asmipl = 0
    pm_total_asym = 0
    pm_total_all = 0
    for (year, month), counts in buckets.items():
        asmipl_c = counts['asmipl']
        asym_c = counts['asym']
        row_total = asmipl_c + asym_c
        pm_total_asmipl += asmipl_c
        pm_total_asym += asym_c
        pm_total_all += row_total
        pm_due_rows.append({
            'month_name': month_names[month - 1],
            'asmipl': asmipl_c,
            'asym': asym_c,
            'total': row_total,
            'year_month': f"{year}-{month:02d}",
        })

    return render(request, "ptr_app/dashboard.html", {
        "designer_workload": designer_workload,
        "total_asmipl": total_asmipl,
        "total_asym": total_asym,
        "grand_total": grand_total,
        "pm_due_rows": pm_due_rows,
        "pm_total_asmipl": pm_total_asmipl,
        "pm_total_asym": pm_total_asym,
        "pm_total_all": pm_total_all,
    })

@login_required
def api_dashboard_stats(request):
    """AJAX JSON Endpoint returning grouped PTR status aggregations in real-time."""
    from django.db.models import Count
    from django.http import JsonResponse
    from .models import PTRRequest
    
    STATUS_MAP = {
        1: "New PTR Request",
        2: "Manager Approved",
        3: "Admin Assigned",
        4: "Designer Submitted",
        5: "Tooling Exec Approved",
        6: "Receiving Completed",
        7: "Fully Active Tools",
        -1: "Rejected Requests",
        -2: "Short Closed",
        -3: "Marked as Common"
    }

    # Pull aggregate counts exclusively tracking natively what lives in DB
    counts = PTRRequest.objects.values('status').annotate(total=Count('status')).order_by('status')
    
    total_ptrs = PTRRequest.objects.count()
    
    # Initialize the results exclusively with zero for uniform frontend consumption securely
    stats = {k: 0 for k in STATUS_MAP.keys()}
    
    # Override zeroes with mathematically pulled truths natively
    for c in counts:
        st = c['status']
        # Unify Designer Short Closed (-4) strictly into Short Closed (-2)
        if st == -4:
            st = -2
            
        if st in stats:
            stats[st] += c['total']
            
    payload = {
        "total": total_ptrs,
        "breakdown": []
    }
    
    # Package into explicit structured payload natively suitable for DOM rendering array securely
    for st, label in STATUS_MAP.items():
        payload["breakdown"].append({
            "status_code": st,
            "label": label,
            "count": stats[st],
            "is_negative": st < 0
        })
        
    return JsonResponse(payload)


# ═══════════════════════════ Tooling Admin ═══════════════════════════


@login_required
@never_cache
def ta_admin_entry_view(request):
    """Render the Tooling Admin - Admin Entry UI."""
    from .forms import MachineModelForm, TypeOfMachineForm
    if request.method == "POST":
        if "machine_model_submit" in request.POST:
            machine_form = MachineModelForm(request.POST)
            type_form = TypeOfMachineForm()
            if machine_form.is_valid():
                machine_form.save()
                return redirect("ta_admin_entry")
        elif "type_of_machine_submit" in request.POST:
            type_form = TypeOfMachineForm(request.POST)
            machine_form = MachineModelForm()
            if type_form.is_valid():
                type_form.save()
                return redirect("ta_admin_entry")
    else:
        machine_form = MachineModelForm()
        type_form = TypeOfMachineForm()

    return render(request, "ptr_app/ta_admin_entry.html", {
        "machine_form": machine_form,
        "type_form": type_form
    })

@login_required
@never_cache
def ta_change_designer_view(request):
    """Tooling Admin - Change Designer: reassign PTRs to a different designer."""
    from .models import PtrAdmin as PtrAdminModel, EmployeeMapping

    if request.method == "POST":
        new_designer = request.POST.get("new_designer", "")
        ptr_ids = request.POST.getlist("ptr_ids")  # list of PtrAdmin PKs

        if not new_designer or not ptr_ids:
            messages.error(request, "Please select PTRs and a new designer.")
            return redirect("ta_change_designer")

        updated = 0
        for pa_id in ptr_ids:
            try:
                pa = PtrAdminModel.objects.get(pk=int(pa_id))
                pa.desginer = new_designer
                # Update dates per-PTR
                new_start = request.POST.get(f"new_start_{pa_id}")
                new_comp = request.POST.get(f"new_comp_{pa_id}")
                if new_start:
                    pa.pldstartdate = new_start
                if new_comp:
                    pa.pldcompdate = new_comp
                pa.save()
                updated += 1
            except PtrAdminModel.DoesNotExist:
                continue

        messages.success(request, f"{updated} PTR(s) reassigned to {new_designer}.")
        return redirect("ta_change_designer")

    return render(request, "ptr_app/ta_change_designer.html")


@login_required
def api_designers_by_unit(request):
    """GET /api/designers-by-unit/?unit=X → JSON list of designers for that unit."""
    from .models import EmployeeMapping
    unit = request.GET.get("unit", "")

    qs = EmployeeMapping.objects.filter(Role="Designer", Status="Active")
    if unit:
        qs = qs.filter(Unit__icontains=unit)

    designers = list(qs.values_list("EmpNumber", flat=True).distinct())
    return JsonResponse({"designers": designers})


@login_required
def api_ptrs_by_designer(request):
    """GET /api/ptrs-by-designer/?designer=X → JSON list of active PTRs for that designer."""
    from .models import PtrAdmin as PtrAdminModel
    designer = request.GET.get("designer", "")
    if not designer:
        return JsonResponse({"ptrs": []})

    pa_qs = PtrAdminModel.objects.filter(
        desginer=designer,
        ptr__status__gte=3
    ).select_related("ptr")

    ptrs = []
    for pa in pa_qs:
        ptrs.append({
            "pa_id": pa.pk,
            "ptrno": pa.ptrno or f"PTR-{pa.ptr.pk:04d}",
            "pldstartdate": pa.pldstartdate.strftime("%d/%m/%Y") if pa.pldstartdate else "",
            "pldcompdate": pa.pldcompdate.strftime("%d/%m/%Y") if pa.pldcompdate else "",
            "pldstartdate_raw": pa.pldstartdate.strftime("%Y-%m-%d") if pa.pldstartdate else "",
            "pldcompdate_raw": pa.pldcompdate.strftime("%Y-%m-%d") if pa.pldcompdate else "",
        })
    return JsonResponse({"ptrs": ptrs})

@login_required
@never_cache
def ta_supplier_resource_view(request):
    """Render the Tooling Admin - Supplier Resource UI."""
    return render(request, "ptr_app/ta_supplier_resource.html")

@login_required
@never_cache
def ta_update_material_view(request):
    """Render the Tooling Admin - Update Material Type UI and handle POST updates."""
    from .models import DesignerEntry
    import json

    selected_unit = request.GET.get("unit", "")

    if request.method == "POST":
        tino = request.POST.get("tino")
        new_material = request.POST.get("new_material")
        frequency = request.POST.get("frequency")
        
        if tino and new_material and frequency:
            de = DesignerEntry.objects.filter(TINo=tino, ptr__status=7).first()
            if de:
                if de.ToolMaterial == new_material:
                    messages.error(request, f"Update rejected: The new material ({new_material}) is exactly the same as the old material.")
                else:
                    de.ToolMaterial = new_material
                    de.PMFrequency = frequency
                    de.save()
                    messages.success(request, f"Successfully updated Material Type and Frequency for TINo: {tino}")
            else:
                messages.error(request, f"TINo: {tino} not found or not eligible for update.")
        else:
            messages.error(request, "Please provide TINo, New Material, and Frequency.")
            
        return redirect(f"{request.path}?unit={selected_unit}")

    # Fetch status 7 entries matching the required conditions
    de_records = DesignerEntry.objects.filter(ptr__status=7).exclude(TINo__isnull=True).exclude(TINo__exact='')
    if selected_unit:
        de_records = de_records.filter(ptr__unit=selected_unit)
    
    tino_data = {}
    for de in de_records:
        tino_data[de.TINo] = {
            "oldMaterial": de.ToolMaterial or "None"
        }
        
    return render(request, "ptr_app/ta_update_material.html", {
        "tino_data": json.dumps(tino_data),
        "de_records": de_records,
        "selected_unit": selected_unit
    })

@login_required
@never_cache
def ta_admin_entry_view_machine(request):
    """Render the Machine Model Data View."""
    machines = MachineModel.objects.all()
    return render(request, "ptr_app/ta_admin_entry_view_machine.html", {"machines": machines})

@login_required
@never_cache
def ta_admin_entry_view_type(request):
    """Render the Type of Machine Data View."""
    types = TypeOfMachine.objects.all()
    return render(request, "ptr_app/ta_admin_entry_view_type.html", {"types": types})

@login_required
@never_cache
def ta_edit_machine_model(request, pk):
    """Edit a Machine Model instance."""
    from .forms import MachineModelForm
    obj = get_object_or_404(MachineModel, pk=pk)
    if request.method == "POST":
        form = MachineModelForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            return redirect("ta_admin_entry_view_machine")
    else:
        form = MachineModelForm(instance=obj)
    return render(request, "ptr_app/ta_edit_machine_model.html", {"form": form, "obj": obj})

@login_required
@never_cache
def ta_edit_type_machine(request, pk):
    """Edit a Type Of Machine instance."""
    from .forms import TypeOfMachineForm
    obj = get_object_or_404(TypeOfMachine, pk=pk)
    if request.method == "POST":
        form = TypeOfMachineForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            return redirect("ta_admin_entry_view_type")
    else:
        form = TypeOfMachineForm(instance=obj)
    return render(request, "ptr_app/ta_edit_type_machine.html", {"form": form, "obj": obj})

@login_required
@never_cache
def ta_delete_machine_model(request, pk):
    """Delete a Machine Model instance."""
    if request.method == "POST":
        obj = get_object_or_404(MachineModel, pk=pk)
        obj.delete()
    return redirect("ta_admin_entry_view_machine")

@login_required
@never_cache
def ta_delete_type_machine(request, pk):
    """Delete a Type Of Machine instance."""
    if request.method == "POST":
        obj = get_object_or_404(TypeOfMachine, pk=pk)
        obj.delete()
    return redirect("ta_admin_entry_view_type")


# ═══════════════════════════ System Admin ═══════════════════════════


@login_required
@never_cache
def sa_design_dashboard_view(request):
    """Render the System Admin - Design Dashboard UI."""
    return render(request, "ptr_app/sa_design_dashboard.html")

@login_required
@never_cache
def sa_employee_mapping_view(request):
    """Render & handle the System Admin - Employee Mapping UI."""
    from .models import EmployeeMapping, ROLE_CHOICES, Department
    from django.utils import timezone

    # All roles from choices
    roles = [r[0] for r in ROLE_CHOICES]

    # All departments for dropdown
    departments = Department.objects.values_list("DeptName", flat=True).distinct()

    # All existing mappings for display
    mappings = EmployeeMapping.objects.all().order_by("-DateTime")

    # Get unique employees for the dropdown
    employees = EmployeeMapping.objects.values_list("EmpNumber", flat=True).distinct()

    if request.method == "POST":
        emp_number = request.POST.get("emp_number", "").strip()
        dept = request.POST.get("department", "").strip()

        if not emp_number:
            messages.error(request, "Please select or enter an employee.")
            return redirect("sa_employee_mapping")

        # Delete existing mappings for this employee (we'll recreate)
        EmployeeMapping.objects.filter(EmpNumber=emp_number).delete()

        # Process each role checkbox
        for role in roles:
            asmipl_checked = request.POST.get(f"role_{role}_ASMIPL") == "on"
            asym_checked = request.POST.get(f"role_{role}_ASYM") == "on"

            if asmipl_checked and asym_checked:
                unit_val = "ASMIPL,ASYM"
            elif asmipl_checked:
                unit_val = "ASMIPL"
            elif asym_checked:
                unit_val = "ASYM"
            else:
                continue  # Neither checked — skip this role

            EmployeeMapping.objects.create(
                EmpNumber=emp_number,
                EmpMail=request.POST.get("emp_mail", ""),
                Status="Active",
                Role=role,
                Unit=unit_val,
                department=dept,
                AddedBy=request.user.username,
            )

        messages.success(request, f"Mapping for {emp_number} saved successfully.")
        return redirect("sa_employee_mapping")

    return render(request, "ptr_app/sa_employee_mapping.html", {
        "roles": roles,
        "departments": departments,
        "mappings": mappings,
        "employees": employees,
    })


@login_required
@never_cache
def sa_employee_mapping_load(request):
    """AJAX: Load existing role-unit mappings for a given employee."""
    from .models import EmployeeMapping
    emp = request.GET.get("emp", "")
    if not emp:
        return JsonResponse({"mappings": [], "department": "", "email": ""})

    qs = EmployeeMapping.objects.filter(EmpNumber=emp)
    mappings = {}
    dept = ""
    email = ""
    for m in qs:
        dept = m.department or ""
        email = m.EmpMail or ""
        units = (m.Unit or "").split(",")
        mappings[m.Role] = units

    return JsonResponse({"mappings": mappings, "department": dept, "email": email})


@login_required
@never_cache
def sa_delete_employee_mapping(request, emp):
    """Delete all mappings for a specific employee."""
    from .models import EmployeeMapping
    if request.method == "POST":
        deleted, _ = EmployeeMapping.objects.filter(EmpNumber=emp).delete()
        messages.success(request, f"Deleted {deleted} mappings for {emp}.")
    return redirect("sa_employee_mapping")


@login_required
@never_cache
def sa_form_mapping_view(request):
    """Render & handle the System Admin - Form Mapping UI."""
    from .models import FormMaster, FormMapping, ROLE_CHOICES, ROLE_ID_MAP

    roles = [r[0] for r in ROLE_CHOICES]
    forms = FormMaster.objects.all().order_by("FormID")

    if request.method == "POST":
        role_name = request.POST.get("role_name", "").strip()
        if not role_name or role_name not in ROLE_ID_MAP:
            messages.error(request, "Please select a valid role.")
            return redirect("sa_form_mapping")

        role_id = ROLE_ID_MAP[role_name]

        # Delete existing mappings for this role
        FormMapping.objects.filter(RolID=role_id).delete()

        # Create new mappings for each checked form
        for form in forms:
            is_checked = request.POST.get(f"form_{form.FormID}") == "on"
            FormMapping.objects.create(
                FormID=form.FormID,
                RolID=role_id,
                IsActive=is_checked,
                AddedBy=request.user.username,
            )

        messages.success(request, f"Form mapping for '{role_name}' saved successfully.")
        return redirect("sa_form_mapping")

    return render(request, "ptr_app/sa_form_mapping.html", {
        "roles": roles,
        "forms": forms,
    })


@login_required
@never_cache
def sa_form_mapping_load(request):
    """AJAX: Load existing form-access mappings for a given role."""
    from .models import FormMapping, ROLE_ID_MAP

    role_name = request.GET.get("role", "")
    if not role_name or role_name not in ROLE_ID_MAP:
        return JsonResponse({"mappings": {}})

    role_id = ROLE_ID_MAP[role_name]
    qs = FormMapping.objects.filter(RolID=role_id, IsActive=True)
    active_forms = {m.FormID: True for m in qs}

    return JsonResponse({"mappings": active_forms})


# ═══════════════════════════ Obsolete ═══════════════════════════


@login_required
@never_cache
def obsolete_fixtures_view(request):
    """Render the Obsolete Fixtures UI & handle submissions."""
    from .models import DesignerEntry, TblRevision, PTRRequest
    
    if request.method == "POST":
        ptr_id = request.POST.get("ptr_id")
        reason = request.POST.get("reason", "")
        remarks = request.POST.get("remarks", "")
        
        if ptr_id:
            ptr = get_object_or_404(PTRRequest, pk=ptr_id)
            # Create the obsolete request pending manager approval
            TblRevision.objects.create(
                ptr=ptr,
                revisionNo=0,
                DescriptionRev=reason,
                RemarksDisposition=remarks,
                isapproved=None
            )
            # Flag PTR as having an obsolete request initialized
            ptr.isobsolete = True
            ptr.save()
            messages.success(request, f"Obsolete request for {ptr.ptrno or f'PTR-{ptr.pk:04d}'} sent for approval.")
            return redirect("obsolete_fixtures")

    selected_unit = request.GET.get("unit", "")

    # Fetch unique TINOs from strictly verified (status 7) PTRs
    entries_qs = DesignerEntry.objects.filter(ptr__status=7, TINo__isnull=False).exclude(TINo="")
    if selected_unit:
        entries_qs = entries_qs.filter(ptr__unit=selected_unit)

    tinos = sorted(list(entries_qs.values_list('TINo', flat=True).distinct()))

    return render(request, "ptr_app/obsolete_fixtures.html", {
        "tinos": tinos,
        "selected_unit": selected_unit
    })


@login_required
def get_ptrs_by_tino(request):
    """AJAX: Fetch active PTR numbers associated with a given TINO."""
    from .models import DesignerEntry
    tino = request.GET.get("tino", "")
    if not tino:
        return JsonResponse({"ptrs": []})
        
    entries = DesignerEntry.objects.filter(TINo=tino, ptr__status=7).select_related('ptr')
    ptrs_data = [{"id": e.ptr.id, "ptrno": e.ptr.ptrno or f"PTR-{e.ptr.id:04d}"} for e in entries]
    
    return JsonResponse({"ptrs": ptrs_data})

# ═══════════════════════════ REPORTS MODULE ═══════════════════════════

@login_required
def reports_view(request):
    """Render the Reports UI with a list of PTRs."""
    ptrs = PTRRequest.objects.all().order_by('-cretedDate')
    return render(request, 'ptr_app/reports.html', {'ptrs': ptrs})

@login_required
def export_report_excel(request, unit):
    """Generate and return the exact Excel formatting based on Screenshots."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from django.http import HttpResponse

    ptrs = PTRRequest.objects.filter(unit=unit).order_by('id')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{unit} Report"
    
    # Define Fills
    fill_main_header = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid") # Dark Green
    fill_grey = PatternFill(start_color="9E9E9E", end_color="9E9E9E", fill_type="solid")
    fill_blue = PatternFill(start_color="90CAF9", end_color="90CAF9", fill_type="solid")
    fill_light_green = PatternFill(start_color="A5D6A7", end_color="A5D6A7", fill_type="solid")
    fill_pink = PatternFill(start_color="F48FB1", end_color="F48FB1", fill_type="solid")
    fill_cyan = PatternFill(start_color="81D4FA", end_color="81D4FA", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")

    # Define Fonts
    font_bold = Font(bold=True, size=10)
    font_header_title = Font(bold=True, size=12)
    font_normal = Font(size=9)
    
    # Define Alignment
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Define Border
    thin = Side(border_style="thin", color="000000")
    border_all = Border(top=thin, left=thin, right=thin, bottom=thin)

    def apply_style(cell, fill=None, font=font_normal, align=align_center):
        if fill: cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border_all
        
    # Set overall width mapping just to make it readable
    for col in range(1, 62):
        col_letter = openpyxl.utils.get_column_letter(col)
        ws.column_dimensions[col_letter].width = 15

    # ──────────────── ROW 1 ────────────────
    ws.merge_cells('A1:BI1')
    cell_r1 = ws['A1']
    if unit == 'ASMIPL':
        cell_r1.value = "AEROSTRUCTURES MANUFACTURING INDIA PVT LTD & AEQUS PVT LTD."
    else:
        cell_r1.value = "AEROSPACE ENGINEERING & MACHINING PVT LTD & AEQUS PVT LTD."
    apply_style(cell_r1, fill_main_header, font_header_title)

    # ──────────────── ROW 2 ────────────────
    ws.merge_cells('A2:L2')
    cell_r2_left = ws['A2']
    cell_r2_left.value = "FORMAT NO: TLG-F017 || REV NO: 04 || EFFECTIVE DATE : 01/APR/2016"
    apply_style(cell_r2_left, fill_main_header, font_header_title, align_left)
    
    ws.merge_cells('M2:BI2')
    cell_r2_right = ws['M2']
    cell_r2_right.value = "COMPLETE REPORT"
    apply_style(cell_r2_right, fill_main_header, font_header_title)
    
    # Apply border to merged cells properly
    for row in ws['A1:BI2']:
        for cell in row:
            cell.border = border_all

    # ──────────────── ROW 3 (Group Headers) ────────────────
    group_headers = [
        ('A3:M3', "PROCESS INPUTS", fill_grey),
        ('N3:W3', "ADMIN", fill_blue),
        ('X3:AL3', "DESIGNING ACTIVITY", fill_light_green),
        ('AM3:AP3', "CM ACTIVITY", fill_pink),
        ('AQ3:AQ3', "FINANCE", fill_light_green),
        ('AR3:AZ3', "TOOL ROOM ACTIVITY", fill_cyan),
        ('BA3:BI3', "PREVENTIVE MAINTENANCE ACTIVITY", fill_yellow),
    ]
    
    for range_str, text, fill in group_headers:
        if ':' in range_str:
            ws.merge_cells(range_str)
            cell = ws[range_str.split(':')[0]]
        else:
            cell = ws[range_str]
            
        cell.value = text
        apply_style(cell, fill, font_bold)
        
        # apply style to merged area
        for row in ws[range_str]:
            for c in row:
                if c.coordinate != cell.coordinate:
                    apply_style(c, fill, font_bold)

    # ──────────────── ROW 4 (Column Headers) ────────────────
    columns = [
        "Sr.No.", "PTR Date & TCMR Date", "Unit", "DEPT", "Customer", "Project", "Part Number", 
        "Part Name", "Part Drawing No.", "OP", "REQ PATH", "Annual QTY", "PTR REMARK",
        "CREATED BY", "APPROVAL MNGR", "MANAGER APPROVED DATE", "MACHINE MODEL", "MACHINE TYPE",
        "Date of Tooling Required", "URGENT REQ REASON", "PTR No & TCMR No.", "Planned Design Start Date",
        "Planned Date for Design Completion", "Designer", "Remarks", "Description of Tooling",
        "Tooling Drawing No.", "REVISION", "Revision Remark", "Actual Date for Design Completion",
        "Reason for Design Delay", "ESTIMATED COST", "Actual Hour", "ROI No.", "Tool Life",
        "Designer Remarks", "TOOLING STATUS", "Updation List of critical spares", "MPR NO", "MPR DATE",
        "CONSUMABLE CODE", "CM Path", "Asset Code", "Suplier", "PO COST", "EXPECTED DELIVERY DATE SCM",
        "GRN date", "GRN number", "Date of verification Tooling", "Verification of Tooling Report Number",
        "Date of tooling Release notice to Production", "Trial Feedback Report No", "PM Frequency in Month",
        "Planned date of Periodic Inspection", "Actual date of Periodic Inspection", "Periodic Inspection Report Number",
        "Periodic Inspection Remarks", "Periodic Inspection Status", "Approved By", "Actual Tool Life",
        "Next due date of Periodic Inspection"
    ]
    
    for idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=4, column=idx)
        cell.value = col_name
        apply_style(cell, font=font_bold)

    # ──────────────── DATA FETCHING ────────────────
    def fmt_date(dt):
        return dt.strftime('%d/%b/%Y') if dt else ""

    for row_idx, ptr in enumerate(ptrs, start=5):
        try:
            dept = Department.objects.get(dID=ptr.deptid).DeptName
        except: dept = ""
        
        try:
            cust = Customer.objects.get(CustID=ptr.customerid).CustomerName
        except: cust = ""
        
        try:
            proj = Project.objects.get(PID=ptr.projectid).prjname
        except: proj = ""

        admin_entry = ptr.admin_entries.first()
        designer_entry = ptr.designer_entries.first()
        receiving_entry = ptr.receiving_entries.first()
        rev_entry = ptr.revisions.order_by('-revisionNo').first()
        pm_entry = ptr.preventative_maintenances.order_by('-actualdate').first()

        data_row = [
            row_idx - 4,                                # A
            fmt_date(ptr.cretedDate),                   # B
            ptr.unit,                                   # C
            dept,                                       # D
            cust,                                       # E
            proj,                                       # F
            ptr.partcode,                               # G
            ptr.partname,                               # H
            ptr.drawing,                                # I
            ptr.process,                                # J
            ptr.req_path,                               # K
            ptr.bqty,                                   # L
            ptr.requestremarks,                         # M
            ptr.createdby,                              # N
            ptr.aprrovedmanager,                        # O
            fmt_date(ptr.mappprovedate),                # P
            ptr.mmodel,                                 # Q
            ptr.typemachine,                            # R
            fmt_date(ptr.dotr),                         # S
            ptr.earlydatereason,                        # T
            ptr.ptrno or "",                            # U
            fmt_date(admin_entry.pldstartdate) if admin_entry else "", # V
            fmt_date(admin_entry.pldcompdate) if admin_entry else "",  # W
            
            # Designer Activity
            admin_entry.desginer if admin_entry else "",             # X
            admin_entry.remarks if admin_entry else "",              # Y
            designer_entry.DescTooling if designer_entry else "",    # Z
            designer_entry.TINo if designer_entry else "",           # AA
            rev_entry.revisionNo if rev_entry else 0,                # AB
            rev_entry.DescriptionRev if rev_entry else "",           # AC
            fmt_date(designer_entry.ACD) if designer_entry else "",  # AD
            designer_entry.latereason if designer_entry else "",     # AE
            designer_entry.ECost if designer_entry else "",          # AF
            designer_entry.actualHr if designer_entry else "",       # AG
            designer_entry.roino if designer_entry else "",          # AH
            designer_entry.toollife if designer_entry else "",       # AI
            designer_entry.desremarks if designer_entry else "",     # AJ
            f"Status {ptr.status}" if ptr.status is not None else "", # AK
            "Yes" if (designer_entry and designer_entry.iscmc) else "No", # AL
            
            # CM Activity
            designer_entry.MPRNo if designer_entry else "",          # AM
            fmt_date(designer_entry.mprDate) if designer_entry else "", # AN
            designer_entry.consumablecode if designer_entry else "", # AO
            designer_entry.cmpath if designer_entry else "",         # AP
            
            # Finance
            designer_entry.erpcode if designer_entry else "",        # AQ 
            
            # Tool Room
            designer_entry.suplier if designer_entry else "",        # AR
            designer_entry.pocost if designer_entry else "",         # AS
            fmt_date(designer_entry.eddate) if designer_entry else "", # AT
            fmt_date(receiving_entry.grndate) if receiving_entry else "", # AU
            receiving_entry.grnno if receiving_entry else "",        # AV
            fmt_date(receiving_entry.rcvdate) if receiving_entry else "", # AW
            receiving_entry.tvrn if receiving_entry else "",         # AX
            "",                                                      # AY (Missing field)
            designer_entry.Tfeedback if designer_entry else "",      # AZ
            
            # PM Activity
            designer_entry.PMFrequency if designer_entry else "",    # BA
            fmt_date(pm_entry.pmdate) if pm_entry else "",           # BB
            fmt_date(pm_entry.actualdate) if pm_entry else "",       # BC
            pm_entry.pmreportno if pm_entry else "",                 # BD
            pm_entry.rmk if pm_entry else "",                        # BE
            pm_entry.status if pm_entry else "",                     # BF
            pm_entry.Approver if pm_entry else "",                   # BG
            "",                                                      # BH (Missing field)
            ""                                                       # BI (Missing field)
        ]

        for idx, val in enumerate(data_row, start=1):
            cell = ws.cell(row=row_idx, column=idx)
            cell.value = val
            apply_style(cell)
            
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Report_{unit}.xlsx"'
    wb.save(response)
    return response
